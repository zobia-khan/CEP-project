#CEP_CS19063



from openpyxl import load_workbook
from random import choice
from string import ascii_uppercase,digits
from datetime import datetime
from abc import ABC, abstractmethod 

#Mix in class
class FileLoader:     
    
    def _loadFile(self,path): 
        try:
           self._wb = load_workbook(path)#To open the workbook,workbook object is created 
           self._sheet_obj = self._wb.active#Get workbook active sheet object from the active attribute
        except:
            print('\nSomething went wrong while loading file')
            
            
            
#Abstract Class            
class user(ABC,FileLoader):
    def __init__(self):
        self.login_status = False
        
    def varifyLogin(self,path):
        
        self._userID=input('\nEnter User ID: ')
        self._password=input('Enter Password: ')
        
        self._loadFile(path)
        
        m_row = self._sheet_obj.max_row #Maximum no of rows of sheet assigned to m_row
        
        for i in range(1, m_row + 1): 
            cell_ID = self._sheet_obj.cell(row = i, column = 1) #cell containing user id
            cell_pass = self._sheet_obj.cell(row = i, column = 2) #cell containing password

            if cell_ID.value == self._userID and cell_pass.value == self._password: #condition: user id and password matches
               self.login_status = True
               
    def get_loginstatus(self):
        return self.login_status
         
    @abstractmethod      
    def createAccount(self):
        pass
    @abstractmethod
    def checkID_inFile(self):
        pass
    @abstractmethod
    def viewAccountInfo(self):
        pass
           



        
class customer(user):
    def __init__(self):
        super().__init__()
        self._customerfp='CEP_CS19063 Customers Data.xlsx' #setting file path
        self.customerAccountInterface()
        
    def customerAccountInterface(self):
        print('\n\n^^^^^^^SIGN-UP OR LOG-IN^^^^^^^^^^^')
        while True:
            print('\n1. LOG IN')
            print('2. SIGN UP')
            print('3. VIEW Account Information')
            print('4. GO to shopping')
            print('press any key to EXIT')
            user_choice = input('Enter your response: ')
            if user_choice == '1':
                self.login()
            elif user_choice == '2':
                self.createAccount()
            elif user_choice == '3':
                self.viewAccountInfo()
            elif user_choice == '4':
                if self.get_loginstatus() == True:
                    self.inst_shoppingCart = shoppingCart(self._userID)
                else:
                    print('\t\tLOG IN to your account first')
                
            else:
                 break
                
        
        
        
    def createAccount(self):
        while True:#infinite loop doesn't break until account is created
            self._userID=input('\nEnter User ID: ')
            
            if self.checkID_inFile(self._userID) == False: #condition for creating a unique user id
    
                self._password=input('Enter Password: ')
                self.__firstName= input('Enter First Name: ')
                self.__lastname= input('Enter Last Name: ')
                self.__email = input('Enter Email Address: ')
                self.__address= input('Enter Current Home Address: ')
                self.__creditCard= input('Enter Credit Card: ')
                
                self._loadFile(self._customerfp)#loading ustomer data file   
                
                self._sheet_obj.append([self._userID,self._password,self.__firstName,self.__lastname,self.__email,self.__address,self.__creditCard])#adding data in a row of excel
                try:
                    self._wb.save(self._customerfp)#saving data to file
                except:
                    print('\nERROR SAVING FILE\nClose CEP_CS19063 Customers Data.xlsx to save data')#catching error when excel file is open
                else:                    
                    print('\n\t\tAccount Created Succesfully')
                    break            
            else:
                print('\nUser name already taken try another')
                
    def login(self):
        
         self.varifyLogin(self._customerfp)   
         if self.login_status == True:
             print('\n\t\tAccount logged in succesfully')
         else:
             print('\nLogin failed')
             
    def viewAccountInfo(self):
        
        self._loadFile(self._customerfp)
        try:
            if self.checkID_inFile(self._userID) == True:
                 print(f'\n{"USER ID":<10}{"PASSWORD":<20}{"FIRST NAME":<20}{"LAST NAME":<20}{"EMAIL ADDRESS":<30}{"HOME ADDRESS":<40}{"CREDIT CARD No"}\n')
                 for value in self._sheet_obj.iter_rows(values_only=True): #in each iteration a list of a row of excel is taken
                    if value[0] == self._userID:#condition: if first element of list matches user id
                        print(f'\n{value[0]:<10}{value[1]:<20}{value[2]:<20}{value[3]:<20}{value[4]:<30}{value[5]:<40}{value[6]}')#printing a user's data
            else:
                 print('\n\nUser ID not Found')
        except:
            print('\n\t\tLOG IN to your account First')
             
             
    def checkID_inFile(self,ID):
        self._loadFile(self._customerfp)
        userIDs=[] #a list of all user IDs in excel
        columnUserID=self._sheet_obj['A']#index A is column 1 in excel containing user IDs
        
        for cell in columnUserID:
            userIDs.append(cell.value) #appending value of cell
        if ID in userIDs:
            return True
        else:
            return False             
             

        
#c1=customer()
#c1.login()
#c1.createAccount()
#c1.viewAccountInfo()
#print(c1.checkID_inFile('john'))



class product(FileLoader):

                
    def displayProducts(self):
        
        self.LoadProductFile()
        print('___________________________________________________________________________')
        print(f'\n\n{"PRODUCT ID":<20}{"PRODUCT NAME":<30}{"UNIT PRICE in Rs"}\n')
        for value in self._sheet_obj.iter_rows(min_row=2,min_col=1,max_col=3,values_only=True):#value is the list of a row values from column 1 to 3
            print(f'\n{value[0]:<20}{value[1]:<30}{value[2]}')#displaying all products information
            
    def getProductDescription(self):
        
        self.LoadProductFile()
        while True:#infinite loop doesn't break until product ID is found in file
            self.pID = input('\nEnter Product ID to get its discription ')#input a product ID from user
            if self.pID in self.getProductIDs(): #checking product ID in file
                pID_row = self.getProductIDs().index(self.pID) #getting index of user input product ID from the list of all products ID
                productName = self._sheet_obj.cell(row = pID_row+1 , column = 2).value #pID_row+1 is the row number of user input product ID
                productinfo = self._sheet_obj.cell(row = pID_row+1 , column = 4).value
                print('\t\t\n\nProduct Name: ',productName,'\n\nDescription: ',productinfo)
                break
                    
            else:
                print('\nProduct ID not found')
                
                
    def LoadProductFile(self):
        
        self._loadFile('CEP_CS19063 Products Data.xlsx')            
            
    def getProductIDs(self):
        
        pdtIDs=[]#list of all product IDs in file
        column_pIDs = self._sheet_obj['A']#column A contains product IDs
        
        for cell in column_pIDs:
            pdtIDs.append(cell.value)
        return pdtIDs    
        

            
#p=product()
#p.displayProducts()
#p.getProductDescription()
#print(p.getProductIDs())



           
class shoppingCart(FileLoader):
    
    def __init__(self,user_id):
        self._userID = user_id
        self._cartitems=[]#nested list of items added to cart 
        self.inst_product = product()
        self.inst_product.LoadProductFile()#loads product file
        self.inst_shoppingHistory = shoppingHistory(user_id)
        self.cartInterface()
    
    def cartInterface(self):
         print('\n\t\t______________________________________________________________________')
         print('\n\t\t^^^^^^^^^^^WELCOME TO ECO-FRIENDLY PRODUCTS STORE^^^^^^^^^^^^^^')
         while True:#breaks when user press any key in user_input
            shoppingCart.displayMenu()
           
            user_choice =input('\nChoose operation from 1 - 7: ' )
                
            if user_choice == '1':
                self.inst_product.displayProducts()
                input()
            elif user_choice == '2':
                self.inst_product.getProductDescription()
            elif user_choice == '3':
                self.addCartItem()
            elif user_choice == '4':
                self.delCartItems()
                input()
            elif user_choice == '5':
                self.viewCartDetails()
                input()
            elif user_choice == '6':
                self.checkOut()
                input()
            elif user_choice == '7':
                self.inst_shoppingHistory.viewShoppingHistory()
                input()
            else:
                print("Are you sure you want to EXIT?(Press any key to exit) ")
                confirm_exit = input()
                if confirm_exit == 'N':
                    pass
                else:
                    break
                    
                
    def displayMenu():
        print('\n\n\t\t\t\t^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^')
        print('\t\t\t\t1. LIST Products' )
        print("\t\t\t\t2. GET a product's description")
        print('\t\t\t\t3. ADD items to cart')
        print('\t\t\t\t4. DELETE items from cart')
        print('\t\t\t\t5. VIEW current cart details')
        print('\t\t\t\t6. CHECKOUT')
        print('\t\t\t\t7. VIEW Shopping History')
        print('\t\t\t\t^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^')
    def addCartItem(self):
        
        while True:
            
            pID = input('\nEnter Product ID to add to cart: ')
                    
            if pID in self.inst_product.getProductIDs():#condition:user input product id in the list of all products IDs

                while True:
                    try:
                        qnty =int(input('Enter product quantity: '))
                        break
                    except:
                        print('\nINVALID INPUT! Enter only integer values')
                pID_row = self.inst_product.getProductIDs().index(pID)#getting index of user input product ID
                product_Quantity = self.inst_product._sheet_obj.cell(row = pID_row+1 , column = 5)#pID + 1 is therow number of user input product ID
                if product_Quantity.value > qnty:#condition:quantity of the product greater than user input quantity
                    
                    productName = self.inst_product._sheet_obj.cell(row = pID_row+1 , column= 2).value
                    productPrice = self.inst_product._sheet_obj.cell(row = pID_row+1 , column = 3).value
                    self._cartitems.append([pID,productName,productPrice,qnty])#Adding the item list in the list of cart items 
                    product_Quantity.value-=qnty#updating stock in file
                    try:
                        self.inst_product._wb.save('CEP_CS19063 Products Data.xlsx')
                    except:
                        print('\nERROR SAVING FILE\nClose CEP_CS19063 Products Data.xlsx to save data')
                        self.addCartItem()
                    else:
                        print(f'\n\t\t{productName} added to cart')
                    while True:
                        user_choice = input("\nWant to add more items? Enter 'Y' for 'YES 'N' for 'NO' ")
                        if user_choice == 'Y':
                            self.addCartItem()
                            break
                        elif user_choice == 'N':
                            break
                        else:
                            print('\nENTER VALID CHOICE')                                         
                else:
                    print('\nSORRY!Item out of stock')
                break   
            else:
                print(f'\nProduct ID "{pID}" not found')
            
    def delCartItems(self):
        if len(self._cartitems) != 0:
            while True:
                pID = input('\nEnter product ID to delete from cart ')
                if pID in self.inst_product.getProductIDs():

                    for item in self._cartitems:
                        if item[0] == pID:#item is the list of item with details,at index zero is the product ID
                            pID_row = self.inst_product.getProductIDs().index(pID)
                            product_Quantity = self.inst_product._sheet_obj.cell(row = pID_row+1 , column = 5)
                            qnty = item[3]#at index 3 is the quantity of product
                            product_Quantity.value+=qnty#updating stock in file
                            try:
                                self.inst_product._wb.save('CEP_CS19063 Products Data.xlsx')
                            except:
                                print('ERROR SAVING FILE\nClose CEP_CS19063 Products Data.xlsx to save data')                            
                            self._cartitems.pop(self._cartitems.index(item))#removing the list of item in cart list
                            print('\n\t\tItem REMOVED successfully') 
                            break
                    else:
                        print('Item not found in cart')
                    break
                else:
                    print('Product ID not found')
        else:
            print('Your Cart is empty')
            
                        
    def viewCartDetails(self):
        if len(self._cartitems) != 0 :#when cart is not empty
            
            print(f'\n\n{"PRODUCT ID":<20}{"PRODUCT NAME":<30}{"PRICE IN Rs.":<30}{"QUANTITY"}\n')
            for item in self._cartitems:
                print(f'\n{item[0]:<20}{item[1]:<30}{item[2]:<30}{item[3]}')
            print(f'\n\t\tTotal Amount {self.calculateTotal()}')
        else:
            print('\t\tYour Cart is empty')
            
            
    def calculateTotal(self):
        total = 0
        for item in self._cartitems:
            total+=item[2]*item[3]#multiplying item's quantity and price
        return total            
                


        
    def checkOut(self):
        print('\n\n^^^^^^^^^^^^^^^^^^^^^^^^^^YOUR CART DETAILS^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^')
        self.viewCartDetails()
        while True:
            user_choice =input("\nDo you want to confirm checkout?\n(Press 'Y' for 'YES' 'N' for 'NO')\nYour choice: ")
            if user_choice == 'y':                
                self.saveCartinFile()#saving shopping history
                self.inst_payment = payment(self._userID)
                try:
                    print(f'CART ID:{self.Cart_id}\nCHECKOUT DATE:{self.checkOutDate}')
                except:
                    print('No cart id generated')
                self.viewCartDetails()
                break
            elif user_choice == 'n':
                print('\nCheckout process CANCELLED')
                break
            else:
                print('ENTER VALID CHOICE!')

    def generateCartKey(self):
        self.Cart_id = ''.join([choice(ascii_uppercase + digits) for n in range(6)]) # a unique string cart ID of lenght 6 
    
    def saveCartinFile(self):
        self._loadFile('CEP_CS19063 Customers Shopping History.xlsx')
        if len(self._cartitems) != 0:
            self.generateCartKey()
            self.checkOutDate = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for item in self._cartitems:
                self._sheet_obj.append([self._userID,self.checkOutDate,self.Cart_id,item[0],item[1],item[2],item[3]])
                try:
                    self._wb.save('CEP_CS19063 Customers Shopping History.xlsx')
                except:
                    print('\n\nERROR SAVING FILE\nClose CEP_CS19063 Customers Shopping History.xlsx to save data')
        else:
            print('Your cart is Empty')        
#s=shoppingCart('zobia')
#s.addCartItem()
#s.delCartItems()
#s.viewCartDetails()
#s.checkOut()


class shoppingHistory(FileLoader):
    def __init__(self,userID):
        self._userID = userID
    def viewShoppingHistory(self):
        self._loadFile('CEP_CS19063 Customers Shopping History.xlsx')
       
        print('\n\n^^^^^^^^^^^^^SHOPPING HISTORY^^^^^^^^^^^^^^^')
        print(f'\n\n{"CHECKOUT DATE":<30}{"CART ID":<10}{"PRODUCT ID":<20}{"PRODUCT NAME":<30}{"PRODUCT PRICE":<20}{"PRODUCT QUANTITY"}\n')
        for value in self._sheet_obj.iter_rows(min_row=2,min_col=1,values_only=True):#value contains list of a row in excel starting from row# 2
           
            if value[0] == self._userID :#user ID is present in cart history
                print(f'{value[1]:<30}{value[2]:<10}{value[3]:<20}{value[4]:<30}{value[5]:<20}{value[6]}')
       
#s= shoppingHistory('hua')
#s.viewShoppingHistory()           
            
                    
class payment(FileLoader):
    def __init__(self,userID):
        self._userID = userID
        self.paymentProcess()
    def paymentProcess(self):
        self._loadFile('CEP_CS19063 Customers Data.xlsx')
        for value in self._sheet_obj.iter_rows(min_row=2,min_col=1,values_only=True):
            if value[0] == self._userID :
                print(f'Credit Card No:{value[6]:<30}')
                while True:
                    user_choice = input('\nUpdate Credit Card No [Y/N]: ')
                    if user_choice == 'y':
                        input('Enter new Credit Card No: ')
                        break
                    elif user_choice == 'n':
                        break
                    else:
                        break
                    
                print('\n\n^^^^^^^^^^^^^^^^^^^^^^^^^^^^INVOICE^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^')
                print('Name:',value[2],value[3])
            

#p=payment('hua')

customer = customer()    
    


    
        

        
        
    
       
        
        
