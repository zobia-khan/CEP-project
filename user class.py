from openpyxl import load_workbook
    
class user:
    def __init__(self):
        self.__loginStatus=False
        self._customerfp='D:\CS-063\OOP\project\sample customer data.xlsx'
        self.fl=FileLoader()
        
    def varifyLogin(self,ID,pasw):
        self._userID=ID
        self._password=pasw
        
        self.fl._loadFile(self._customerfp)
        
        m_row = self.fl._sheet_obj.max_row 
        for i in range(1, m_row + 1): 
            cell_ID = self.fl._sheet_obj.cell(row = i, column = 1)
            cell_pass = self.fl._sheet_obj.cell(row = i, column = 2)
            if cell_ID.value == self._userID and cell_pass.value == self._password:
               self.__loginStatus = True
               return self.__loginStatus
           
class FileLoader:      
    def _loadFile(self,path): 
        try:
           self._wb = load_workbook(path)
           self._sheet_obj = self._wb.active
        except:
            return
#u1=user()
#u1.varifyLogin('user id','password')
#print(u1.loginStatus)
        
class customer(user):
        
    def createAccount(self,ID,pasw,fn,ln,email,adrs,cno):
        self._userID= ID
        self._password= pasw
        self.__firstName= fn
        self.__lastname= ln
        self.__email = email
        self.__address= adrs
        self.__creditCard= cno
        self.fl._loadFile(self._customerfp)
        userIDs=[]
        columnUserID=self.fl._sheet_obj['A']
        for cell in columnUserID:
            userIDs.append(cell.value)
            
        if ID not in userIDs:
            self.fl._sheet_obj.append([self._userID,self._password,self.__firstName,self.__lastname,self.__email,self.__address,self.__creditCard])
            self.fl._wb.save(self._customerfp)
            return 'Account created succesfully'
        else:
            return 'User name already taken try another'
#    def updateAccount(self,ID):
#       self._filereader('D:\CS-063\OOP\project\sample customer data.xlsx')
#        userIDs=[]
#        m_row = self._sheet_obj.max_row
#        for i in range(1, m_row + 1): 
#            userid = self._sheet_obj.cell(row = i, column = 1) 
#            userIDs.append(userid.value)
#        if ID in userIDs:
#            self.__userID=ID
        
#c1=customer()
#print(c1.createAccount('ue id','pasdfs','lion','king','lionking@yahoo.com','24321/342 sfv','visa 2144325'))

class product:
    def LoadProductFile(self):
        self.fl = FileLoader()
        self.fl._loadFile('D:\CS-063\OOP\project\products data.xlsx')
    def getProductIDs(self):
        pdtIDs=[]
        column_pIDs = self.fl._sheet_obj['A']
        for cell in column_pIDs:
            pdtIDs.append(cell.value)
        return pdtIDs
        
    def displayProducts(self):
        self.LoadProductFile()
        print(f'\n\n{"PRODUCT ID":<20}{"PRODUCT NAME":<30}{"PRICE in Rs"}\n')
        for value in self.fl._sheet_obj.iter_rows(min_row=2,min_col=1,max_col=3,values_only=True):
            print(f'\n{value[0]:<20}{value[1]:<30}{value[2]}')
            
    def getProductDescription(self,pID):
        self.LoadProductFile()
        self.getProductIDs()
        if pID in self.getProductIDs(): 
            productName = self.fl._sheet_obj.cell(row = pID+1 , column = 2).value
            productinfo = self.fl._sheet_obj.cell(row = pID+1 , column = 4).value
            print('\n\nProduct Name: '+productName+'\n\nDescription: '+productinfo)
        else:
            print('\nProduct ID not found')


#p=product()
#p.displayProducts()
#p.getProductDescription(4)

class shoppingCart:
    def __init__(self):
        self._cartitems=[]
        self.inst_product = product()
        self.inst_product.LoadProductFile()
        
    def addCartItem(self,pID,qnty):
        if pID in self.inst_product.getProductIDs():           
            productPrice = self.inst_product.fl._sheet_obj.cell(row = pID+1 , column = 3).value
            self._cartitems.append([pID,productPrice,qnty])
            print('\nItem added to cart')
        else:
            print(f'\nProduct ID "{pID}" not found')
            
    def delCartItems(self,pID):
        if pID in self.inst_product.getProductIDs():
            for item in self._cartitems:
                if item[0] == pID:
                    self._cartitems.pop(self._cartitems.index(item))
                    print('Item removed successfully')
                else:
                    print('Item not found in cart')
        else:
            print('Product ID not found')
            
    def updateQuantity(self,pID,qty):
        if pID in self.inst_product.getProductIDs():
            for item in self._cartitems:
                if item[0] == pID:
                    self._cartitems[self._cartitems.index(item)][item[2]] = qty
                    print(f'Item quantity updated to "{qty}"')
                else:
                    print('Item not found in cart')
        else:
            print(f'Product ID "{pID}" not found')
    def viewCartDetails(self):
        print(f'\n\n{"PRODUCT ID":<20}{"PRICE in Rs":<30}{"QUANTITY"}\n')
        for item in self._cartitems:
            print(f'\n{item[0]:<20}{item[1]:<30}{item[2]}')
                    
            
s=shoppingCart()
s.addCartItem(5,1)
s.delCartItems(6)
s.delCartItems(6)
s.updateQuantity(0,4)
s.viewCartDetails()
        
        

        
        
    
       
        
        